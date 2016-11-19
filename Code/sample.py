'Module to get Wikipedia entry information'

_author_ = "Yidao@Babacheku"

import time,os

from openpyxl.styles import PatternFill,Alignment,Font
from openpyxl.styles.colors import RED,WHITE,BLUE,DARKYELLOW
                                                                                        # 导入网页自动化测试工具selenium模块
from selenium.common.exceptions import NoSuchElementException,TimeoutException,StaleElementReferenceException         # 导入异常模块

os.environ["SELENIUM_SERVER_JAR"] = "/Users/bin/Desktop/SourceTree/BaikeEnv/selenium/selenium-server-standalone-2.48.0.jar"   # 添加selenium服务器地址环境变量

def Data_Scratch( browser ):

    # results[0]: 百度词条名称
    # results[1]: 概述字数
    # results[2]: 基本信息栏条数
    # results[3]: 一级目录个数
    # results[4]: 正文字数
    # results[5]: 参考文献条数
    # results[6]: 图片条数
    # results[7]: "科普中国百科"是否收录

    results = []

    # 获取特征点
    # 获取百科词条名称
    try:
        print("找到词条名称")
        results.append(browser.find_element_by_tag_name("h1").text)
    except NoSuchElementException:
        results.append("None")

    print("百科词条名称：%s" % results[-1])

    # 获取概述段落的字数
    try:
        browser.find_element_by_xpath("//div[@label-module='lemmaSummary']")
        results.append(len(browser.find_element_by_xpath("//div[@label-module='lemmaSummary']").text))
    except NoSuchElementException:
        results.append(-1)

    print("概述字数:%d" % results[-1])  # debug

    # 获取基本信息栏条数
    try:
        browser.find_element_by_xpath("//dt[@class='basicInfo-item name']")
        print("找到基本信息栏")
        results.append(len(browser.find_elements_by_xpath("//dt[@class='basicInfo-item name']")))
    except NoSuchElementException:
        results.append(-1)

    print("基本信息栏条数：%d" % results[-1])  # debug

    # 获取一级目录条数
    try:
        browser.find_element_by_xpath("//div[@class='lemma-catalog']/div/ol/li[@class='level1']")
        print("找到一级目录")
        results.append(len(browser.find_elements_by_xpath("//div[@class='lemma-catalog']/div/ol/li[@class='level1']")))
    except NoSuchElementException:
        results.append(-1)

    print("一级目录条数:%d" % results[-1])  # debug

    # 获取正文段数及字数
    try:
        browser.find_element_by_xpath("//div[@label-module='para']")
        Content_Paragraphs = browser.find_elements_by_xpath("//div[@label-module='para']")

        Number_Of_Words = 0
        for Content_Paragraph in Content_Paragraphs:
            Number_Of_Words += len(Content_Paragraph.text)

        results.append(Number_Of_Words)
    except NoSuchElementException:
        results.append(-1)

    print("正文字数:%d" % results[-1])  # debug

    # 获取参考文献条数
    try:
        browser.find_element_by_xpath("//ul[@class='reference-list']/li[@class='reference-item ']")
        print("找到参考文献")
        results.append(
            len(browser.find_elements_by_xpath("//ul[@class='reference-list']/li[@class='reference-item ']")) + len(
                browser.find_elements_by_xpath("//ul[@class='reference-list']/li[@class='reference-item more']")))
    except NoSuchElementException:
        results.append(-1)

    print("参考文献条数:%d" % results[-1])

    # 获取词条图片数量
    try:
        browser.find_element_by_link_text("更多图册")
        print("找到多个图册")
        browser.get(browser.find_element_by_link_text("更多图册").get_attribute('href'))
        browser.implicitly_wait(1)

        results.append(int(browser.find_element_by_xpath("//span[@class='pic-num num']").text))

        browser.get(browser.find_element_by_class_name("return-back").get_attribute('href'))
        browser.implicitly_wait(1)
    except NoSuchElementException:
        try:
            browser.find_element_by_class_name("summary-pic")
            print("找到一个图册")
            browser.get(browser.find_element_by_xpath("//div[@class='summary-pic']/a").get_attribute('href'))
            browser.implicitly_wait(1)

            results.append(int(browser.find_element_by_xpath("//span[@style='color:#427cb8']").text))

            browser.get(browser.find_element_by_link_text("返回词条").get_attribute('href'))
            browser.implicitly_wait(1)
        except NoSuchElementException:
            results.append(-1)

    print("词条图片数量:%d" % results[-1])

    return results

def Excellent_Label( broweser ):

    # 检查词条是否为"特色词条"
    try:
        broweser.find_element_by_xpath("//a[@class='posterFlag excellent-icon']")
        return True
    except NoSuchElementException:
        return False

def Included_Label( browser ):

    # 查询词条是否被"科普中国百科"收录

    try:
        browser.find_element_by_class_name("professional-con")
        result = "已收录"
    except NoSuchElementException:
        try:
            browser.find_element_by_id("authEdit")
            result = "已收录"
        except NoSuchElementException:
            try:
                browser.find_element_by_id("authResource")
                result = "已收录"
            except NoSuchElementException:
                result = "未收录"

    return result

def Get_Entry_Page( keyword,browser,keylist ):

    # -1: 百度百科未收录
    # 百科词条Safari句柄

    print("Python Selenium Safari Started")  # 程序开始运行提示

    while True:
        browser.get("http://baike.baidu.com/")  # 打开网址

        try:
            browser.find_element_by_xpath("//div[@class='logo wiki-home-slogan']")
            break
        except NoSuchElementException:
            continue

    browser.implicitly_wait(2)

    baike_search_key = browser.find_element_by_id("query")  # 按网页元素id查找网页元素query
    baike_search_key.clear()  # 清除输入框里的内容
    baike_search_key.send_keys(keyword)  # 将获取的数据添加到输入框里

    browser.find_element_by_id("search").click()  # 单击搜索按钮

    browser.implicitly_wait(2)

    # 查询词条是否被"百度百科"收录
    while True:
        try:
            browser.find_element_by_class_name("create-entrance")

            # 若"百度百科首页"查询未收录，查询"百度首页"
            while True:
                browser.get("http://wwww.baidu.com/")
                try:
                    browser.find_element_by_id("su")
                    break
                except NoSuchElementException:
                    continue

            browser.implicitly_wait(2)

            baidu_search_key = browser.find_element_by_id("kw")
            baidu_search_key.clear()
            baidu_search_key.send_keys(keyword)

            browser.find_element_by_id("su").click()
            browser.implicitly_wait(2)

            try:
                browser.find_element_by_partial_link_text("百度百科")

                # 获取搜索结果页中包含"百度百科"字样的链接
                Search_Results = browser.find_elements_by_partial_link_text("百度百科")

                # 搜索结果验证
                for Search_Result in Search_Results:

                    if keyword in Search_Result.text:
                        continue
                    else:
                        Search_Results.remove(Search_Result)

                    print(Search_Result.text)

                print(len(Search_Results))  # debug

                if len(Search_Results) > 0 and keyword in Search_Results[0].text:

                    Switch_Link = Search_Results[0].get_attribute('href')
                    print("词条链接为：%s" % (Switch_Link))  # Debug

                    # 打开词条页面
                    browser.get(Switch_Link)
                    browser.implicitly_wait(2)

                    return browser
                else:
                    raise NoSuchElementException
            except NoSuchElementException:

                return -1
        except NoSuchElementException:

            # 查询是否是多义词
            try:
                browser.find_element_by_class_name("lemmaWgt-subLemmaListTitle")

                Poly_Entries = browser.find_elements_by_partial_link_text(keyword)

                print("义项个数：%d" % (len(Poly_Entries)))

                Entries_Valid = []

                for Entry in Poly_Entries:

                    Entry_Text = Entry.text

                    for Key in keylist:
                        if Key in Entry_Text:
                            Entries_Valid.append(Entry)
                            break
                        else:
                            continue

                Entries_Valid.append(Poly_Entries[0])

                if len(Entries_Valid) == 1:
                    print("只有一个义项符合要求，正确")
                else:
                    print("有多个义项符合要求，错误")

                if Entries_Valid[0].get_attribute('href'):
                    Valid_Entry_Link = Entries_Valid[0].get_attribute('href')

                    print("符合要求义项链接：%s" % (Valid_Entry_Link))
                    browser.get(Valid_Entry_Link)
                    browser.implicitly_wait(2)

                return browser
            except NoSuchElementException:
                try:
                    browser.find_element_by_class_name("polysemantList-header-title")

                    while True:
                        browser.get(browser.find_element_by_partial_link_text("个义项").get_attribute('href'))

                        try:
                            browser.find_element_by_class_name("lemmaWgt-subLemmaListTitle")
                            break
                        except NoSuchElementException:
                            continue

                    Poly_Entries = browser.find_elements_by_partial_link_text(keyword)

                    if len(Poly_Entries) == 0:
                        Poly_Entries = browser.find_elements_by_xpath("//div[@label-module='para']/a")

                    print("义项个数：%d" % (len(Poly_Entries)))

                    Entries_Valid = []

                    for Entry in Poly_Entries:

                        Entry_Text = Entry.text

                        for Key in keylist:
                            if Key in Entry_Text:
                                Entries_Valid.append(Entry)
                                break
                            else:
                                continue

                    Entries_Valid.append(Poly_Entries[0])

                    if len(Entries_Valid) == 1:
                        print("只有一个义项符合要求，正确")
                    else:
                        print("有多个义项符合要求，错误")

                    if Entries_Valid[0].get_attribute('href'):
                        Valid_Entry_Link = Entries_Valid[0].get_attribute('href')

                        print("符合要求义项链接：%s" % (Valid_Entry_Link))
                        browser.get(Valid_Entry_Link)
                        browser.implicitly_wait(2)

                    return browser
                except NoSuchElementException:
                    print("本词条非多义词")
                    return browser
        except TimeoutException:
            continue
        except StaleElementReferenceException:
            continue

class MarkFormat:

    # 表格格式调整
    Error_Fill = PatternFill(patternType='solid', start_color=RED, end_color=RED)
    Not_Included_Fill = PatternFill(patternType='solid', start_color=BLUE, end_color=BLUE)
    Excellent_Included_Fill = PatternFill(patternType='solid',start_color=DARKYELLOW,end_color=DARKYELLOW)
    Error_Font = Font(color=WHITE)

class KeyList:

    # 词表关键字
    Aero_Avia_Keys = ['航空', '航天', '飞行器', '飞机', '导弹', '客机', '战斗机', '轰炸机', '歼击机', '攻击机', '运输机', '直升机','无人机',
                      '火箭', '卫星', '空间站', '探测器', '飞船', '宇宙', '地球', '月球', '太阳', '深空', '火星','战机','武器']

    Astronomy_Keys = ['天文','天文台','望远镜']

    Climate_Keys = ['自然','大气','气象','天气','气候','词语概念']

    Information_Keys = ['软件','程序','计算机','网络','互联网','协议','链接','代码','编程','网页','通信','通讯','传讯','磁盘','漏洞','自动化','控制']

    Energy_Keys = ['能源','材料','电池']