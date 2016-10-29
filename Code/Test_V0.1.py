from selenium import webdriver                                                         # 导入网页自动化测试工具selenium模块
from selenium.common.exceptions import NoSuchElementException,TimeoutException,StaleElementReferenceException         # 导入异常模块

from openpyxl import load_workbook                                                     # 导入Excel文件读取模块

from openpyxl.styles import PatternFill,Alignment,Font
from openpyxl.styles.colors import RED,WHITE,BLUE
from WikiSpider.DataScratch import Baike_Data_Scratch

import time                                                                            # 导入时间模块
import os                                                                              # 导入操作系统模块

os.environ["SELENIUM_SERVER_JAR"] = "/Users/bin/Desktop/SourceTree/BaikeEnv/selenium/selenium-server-standalone-2.48.0.jar"   # 添加selenium服务器地址环境变量

print("Python Selenium Safari Started")                       # 程序开始运行提示

safari = webdriver.Safari()                                   # 打开safari浏览器

WB = load_workbook('/Users/bin/Desktop/InProgress/词表处理/临时词表/航天词表_全部.xlsx')                  # 打开Excel文件
#WB = load_workbook('test_Poly.xlsx')

# 表格格式调整
Error_Fill = PatternFill(patternType='solid',start_color= RED,end_color=RED)
Not_Included_Fill = PatternFill(patternType='solid',start_color=BLUE,end_color=BLUE)
Cell_Aligment = Alignment(horizontal='center',vertical='center')
Error_Font = Font(color=WHITE)

Aero_Avia_Keys = ['航空','航天','飞行器','飞机','导弹','客机','战斗机','轰炸机','歼击机','攻击机','运输机','直升机','无人机',
                  '火箭','卫星','空间站','探测器','飞船','宇宙','地球','月球','太阳','深空','火星']

Key_List = Aero_Avia_Keys

# 主循环，遍历Worksheets
for ws_index in range(len(WB.sheetnames)):

    WS = WB.worksheets[ws_index]                                         # 打开worksheetNo

    Head_Column_No = WS.max_column

    print("表格表头共%d行，%d列" % (WS.max_row, WS.max_column))            # debug

    # 判断词条名称所在列
    for item_index in range(1,(Head_Column_No + 1)):
        if WS.cell(row = 2, column = item_index).value == "词条名称":
            Entry_Column_Index = item_index

    print(Entry_Column_Index)

    WS.cell(row = 2, column = Head_Column_No + 1).value = "百科词条名"
    WS.cell(row = 2, column = Head_Column_No + 2).value = "是否被\"百度百科\"收录"
    WS.cell(row = 2, column = Head_Column_No + 3).value = "是否被\"科普中国百科\"收录"
    WS.cell(row = 2, column = Head_Column_No + 4).value = "词条等级"
    WS.cell(row = 2, column = Head_Column_No + 5).value = "词条网址"

    WS.cell(row = 2, column = Head_Column_No + 6).value = "概述字数"
    WS.cell(row = 2, column = Head_Column_No + 7).value = "基本信息栏条数"
    WS.cell(row = 2, column = Head_Column_No + 8).value = "一级目录条数"
    #WS.cell(row = 2, column = Head_Column_No + 9).value = "二级目录条数"
    #WS.cell(row = 2, column = Head_Column_No + 10).value = "正文段数"
    WS.cell(row = 2, column = Head_Column_No + 9).value = "正文字数"
    WS.cell(row = 2, column = Head_Column_No + 10).value = "参考文献条数"
    #WS.cell(row = 2, column = Head_Column_No + 13).value = "词条图册数"
    WS.cell(row = 2, column = Head_Column_No + 11).value = "词条图片张数"

    Row_Count = 1                                # 统计已完成词条数

    for row_index in range(3,(WS.max_row + 1)):

        KeyWord = WS.cell(row = row_index,column = Entry_Column_Index).value                 # 获取当前sheet第i行，词条名称列数据值并打印

        if KeyWord == "":
            continue

        while True:
            safari.get("http://baike.baidu.com/")                               # 打开网址

            try:
                safari.find_element_by_xpath("//div[@class='logo wiki-home-slogan']")
                break
            except NoSuchElementException:
                continue

        safari.implicitly_wait(2)

        baike_search_key = safari.find_element_by_id("query")               # 按网页元素id查找网页元素query
        baike_search_key.clear()                                            # 清除输入框里的内容
        baike_search_key.send_keys(KeyWord)                                 # 将获取的数据添加到输入框里

        safari.find_element_by_id("search").click()                         # 单击搜索按钮

        safari.implicitly_wait(2)

        #查询词条是否被"百度百科"收录

        Timeout_Flag = 1

        while Timeout_Flag == 1:
            try:
                safari.find_element_by_class_name("create-entrance")

                # 若"百度百科首页"查询未收录，查询"百度首页"
                while True:
                    safari.get("http://wwww.baidu.com/")
                    try:
                        safari.find_element_by_id("su")
                        break
                    except NoSuchElementException:
                        continue

                safari.implicitly_wait(2)

                baidu_search_key = safari.find_element_by_id("kw")
                baidu_search_key.clear()
                baidu_search_key.send_keys(KeyWord)

                safari.find_element_by_id("su").click()
                safari.implicitly_wait(2)

                try:
                    safari.find_element_by_partial_link_text("百度百科")

                    # 获取搜索结果页中包含"百度百科"字样的链接
                    Search_Results = safari.find_elements_by_partial_link_text("百度百科")

                    # 搜索结果验证
                    for Search_Result in Search_Results:

                        if KeyWord in Search_Result.text:
                            continue
                        else:
                            Search_Results.remove(Search_Result)

                        print(Search_Result.text)

                    print(len(Search_Results))               # debug


                    if len(Search_Results) > 0 and KeyWord in Search_Results[0].text:
                        WS.cell(row = row_index, column = Head_Column_No + 2).value = "已收录"

                        Switch_Link = Search_Results[0].get_attribute('href')
                        print("词条链接为：%s" %(Switch_Link))                                     # Debug

                        # 打开词条页面
                        safari.get(Switch_Link)
                        safari.implicitly_wait(2)

                        WS.cell(row=row_index, column=Head_Column_No + 5).value = safari.current_url

                        # 获取词条特征数据
                        Entry_Data = Baike_Data_Scratch(safari)
                    else:
                        raise NoSuchElementException
                except NoSuchElementException:
                    WS.cell(row = row_index, column = Head_Column_No + 2).value = "未收录"
                    WS.cell(row = row_index, column = Head_Column_No + 5).value = "None"

                    Entry_Data = []
                    Entry_Data.append("None")
                    for index in range(1,8):
                        Entry_Data.append(-1)

                Timeout_Flag = 0
            except NoSuchElementException:
                WS.cell(row = row_index, column = Head_Column_No + 2).value = "已收录"

                # 查询是否是多义词
                try:
                    safari.find_element_by_class_name("lemmaWgt-subLemmaListTitle")

                    Poly_Entries = safari.find_elements_by_partial_link_text(KeyWord)

                    print("义项个数：%d" % (len(Poly_Entries)))

                    Entries_Valid = []

                    for Entry in Poly_Entries:

                        Entry_Text = Entry.text

                        for Key in Key_List:
                            if Key in Entry_Text:
                                Entries_Valid.append(Entry)
                                break
                            else:
                                continue

                    Entries_Valid.append(Poly_Entries[0])

                    if len(Entries_Valid) == 1:
                        print("只有一个义项符合要求，正确")
                    else:
                        print("有多个义项符合要求，错误")

                    if Entries_Valid[0].get_attribute('href'):
                        Valid_Entry_Link = Entries_Valid[0].get_attribute('href')

                        print("符合要求义项链接：%s" % (Valid_Entry_Link))
                        safari.get(Valid_Entry_Link)
                        safari.implicitly_wait(2)

                    WS.cell(row=row_index, column=Head_Column_No + 5).value = safari.current_url
                except NoSuchElementException:
                    try:
                        safari.find_element_by_class_name("polysemantList-header-title")

                        while True:
                            safari.get(safari.find_element_by_partial_link_text("个义项").get_attribute('href'))

                            try:
                                safari.find_element_by_class_name("lemmaWgt-subLemmaListTitle")
                                break
                            except NoSuchElementException:
                                continue


                        Poly_Entries = safari.find_elements_by_partial_link_text(KeyWord)

                        if len(Poly_Entries) == 0:
                            Poly_Entries = safari.find_elements_by_xpath("//div[@label-module='para']/a")

                        #Poly_Entries = safari.find_elements_by_xpath("//ul/li[@class='item']/a")
                        #Poly_Entries.insert(0,safari.find_element_by_xpath("//ul/li[@class='item']/span[@class='selected']"))

                        print("义项个数：%d" % (len(Poly_Entries)))

                        Entries_Valid = []

                        for Entry in Poly_Entries:

                            Entry_Text = Entry.text

                            for Key in Key_List:
                                if Key in Entry_Text:
                                    Entries_Valid.append(Entry)
                                    break
                                else:
                                    continue

                        Entries_Valid.append(Poly_Entries[0])

                        if len(Entries_Valid) == 1:
                            print("只有一个义项符合要求，正确")
                        else:
                            print("有多个义项符合要求，错误")

                        if Entries_Valid[0].get_attribute('href'):

                            Valid_Entry_Link = Entries_Valid[0].get_attribute('href')

                            print("符合要求义项链接：%s" % (Valid_Entry_Link))
                            safari.get(Valid_Entry_Link)
                            safari.implicitly_wait(2)

                        WS.cell(row=row_index, column=Head_Column_No + 5).value = safari.current_url
                    except NoSuchElementException:
                        print("本词条非多义词")
                        WS.cell(row=row_index, column=Head_Column_No + 5).value = safari.current_url

                # 获取特征点
                Entry_Data = Baike_Data_Scratch(safari)

                Timeout_Flag = 0
            except TimeoutException:
                Timeout_Flag = 1
            except StaleElementReferenceException:
                Timeout_Flag = 1

        WS.cell(row = row_index, column = Head_Column_No + 1).value = Entry_Data[0]           # 百度百科词条名称
        WS.cell(row = row_index, column = Head_Column_No + 3).value = Entry_Data[-1]          # "科普中国百科"是否收录

        for Result_index in range(1,(len(Entry_Data) - 1)):
            WS.cell(row = row_index, column = Head_Column_No + 5 + Result_index).value = Entry_Data[Result_index]

        # 评价函数
        if WS.cell(row = row_index, column = Head_Column_No + 1).value != WS.cell(row = row_index, column = Entry_Column_Index).value:
            WS.cell(row = row_index, column = Head_Column_No + 1).fill = Error_Fill
            WS.cell(row = row_index, column = Head_Column_No + 1).font = Error_Font

        # 百度百科未收录
        if WS.cell(row = row_index, column = Head_Column_No + 2).value == "未收录":                     # 百度百科未收录的为新建词条
            WS.cell(row = row_index, column = Head_Column_No + 2).font = Error_Font
            WS.cell(row = row_index, column = Head_Column_No + 2).fill = Not_Included_Fill

            WS.cell(row = row_index, column = Head_Column_No + 4).value = 4                      # 百度百科未收录词条，直接置为第四等级，需新建
        elif WS.cell(row = row_index, column = Head_Column_No + 3).value == "已收录":             # "科普中国百科"已收录的词条不需要建设
            WS.cell(row = row_index, column = Head_Column_No + 3).font = Error_Font
            WS.cell(row = row_index, column = Head_Column_No + 3).fill = Error_Fill
            WS.cell(row = row_index, column = Head_Column_No + 4).value = -1
        else:                                                                                     # 百度百科已收录而"科普中国百科"未收录
            if Entry_Data[4] == -1:                                                               # 词条中没有正文，级别4
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 4
            elif Entry_Data[4] <= 1000:                                                           # 正文内容过少或缺少目录，级别3
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 3
            elif Entry_Data[3] <= 0:
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 3
            elif Entry_Data[2] <= 0:                                                             # 缺乏基本信息栏或参考文献，级别2
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 2
            elif Entry_Data[5] <= 0:
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 2
            elif Entry_Data[1] <= 0:                                                             # 缺乏概述或图片，级别2
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 2
            elif Entry_Data[6] <= 0:
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 2
            else:                                                                                 # 无结构性缺失，内容不过少，等级1
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 1

        Row_Count += 1

        WB.save('/Users/bin/Desktop/InProgress/词表处理/输出词表/航天词表_全部_提交.xlsx')
        #WB.save('output.xlsx')
        # 每20行保存一次
        #if Row_Count == 5:
            #WB.save('/Users/bin/Desktop/InProgress/词表处理/输出词表/临时输出.xlsx')
        #else:
           # continue

        #time.sleep(2)

    #WB.save('/Users/bin/Desktop/InProgress/词表处理/输出词表/临时输出.xlsx')                                   # 保存输出的Excel文件



safari.close()                                                        # 关闭safari窗口

safari.quit()                                                         # 推出safari
