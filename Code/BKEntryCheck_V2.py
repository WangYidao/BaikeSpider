from selenium import webdriver                                                         # 导入网页自动化测试工具selenium模块
from selenium.common.exceptions import NoSuchElementException                          # 导入异常模块

from openpyxl import load_workbook                                                     # 导入Excel文件读取模块

from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
from WikiSpider.DataScratch import Baike_Data_Scratch

import time                                                                            # 导入时间模块
import os                                                                              # 导入操作系统模块

os.environ["SELENIUM_SERVER_JAR"] = "/Users/bin/Desktop/SourceTree/BaikeEnv/selenium/selenium-server-standalone-2.48.0.jar"   # 添加selenium服务器地址环境变量

print("Python Selenium Safari Started")                       # 程序开始运行提示

safari = webdriver.Safari()                                   # 打开safari浏览器

WB = load_workbook('test_Poly.xlsx')                  # 打开Excel文件

Aero_Avia_Keys = ['航空','航天','飞行器','飞机','导弹','客机','战斗机','轰炸机','歼击机','攻击机','运输机','直升机','无人机',
                  '火箭','卫星','空间站','探测器','飞船']

Key_List = Aero_Avia_Keys

# 主循环，遍历Worksheets
for ws_index in range(len(WB.sheetnames)):

    WS = WB.worksheets[ws_index]                                         # 打开worksheetNo

    Head_Column_No = WS.max_column

    print("表格表头共%d行，%d列" % (WS.max_row, WS.max_column))            # debug

    # 判断词条名称所在列
    for item_index in range(1,(Head_Column_No + 1)):
        if WS.cell(row = 1, column = item_index).value == "词条名称":
            Entry_Column_Index = item_index

    print(Entry_Column_Index)

    WS.cell(row = 1, column = Head_Column_No + 1).value = "百科词条名"
    WS.cell(row = 1, column = Head_Column_No + 2).value = "是否被\"百度百科\"收录"
    WS.cell(row = 1, column = Head_Column_No + 3).value = "是否被\"科普中国百科\"收录"
    WS.cell(row = 1, column = Head_Column_No + 4).value = "词条网址"

    WS.cell(row = 1, column = Head_Column_No + 5).value = "概述字数"
    WS.cell(row = 1, column = Head_Column_No + 6).value = "基本信息栏条数"
    WS.cell(row = 1, column = Head_Column_No + 7).value = "一级目录条数"
    WS.cell(row = 1, column = Head_Column_No + 8).value = "二级目录条数"
    WS.cell(row = 1, column = Head_Column_No + 9).value = "正文段数"
    WS.cell(row = 1, column = Head_Column_No + 10).value = "正文字数"
    WS.cell(row = 1, column = Head_Column_No + 11).value = "参考文献条数"
    WS.cell(row = 1, column = Head_Column_No + 12).value = "词条图册数"
    WS.cell(row = 1, column = Head_Column_No + 13).value = "词条图片张数"

    for row_index in range(2,(WS.max_row + 1)):

        KeyWord = WS.cell(row = row_index,column = Entry_Column_Index).value                 # 获取当前sheet第i行，词条名称列数据值并打印

        safari.get("http://baike.baidu.com/")                               # 打开网址
        safari.implicitly_wait(2)

        baike_search_key = safari.find_element_by_id("query")               # 按网页元素id查找网页元素query
        baike_search_key.clear()                                            # 清除输入框里的内容
        baike_search_key.send_keys(KeyWord)                                 # 将获取的数据添加到输入框里

        safari.find_element_by_id("search").click()                         # 单击搜索按钮

        safari.implicitly_wait(2)

        #查询词条是否被"百度百科"收录
        try:
            safari.find_element_by_class_name("create-entrance")

            # 若"百度百科首页"查询未收录，查询"百度首页"
            safari.get("http://wwww.baidu.com/")
            safari.implicitly_wait(2)

            baidu_search_key = safari.find_element_by_id("kw")
            baidu_search_key.clear()
            baidu_search_key.send_keys(KeyWord)

            safari.find_element_by_id("su").click()
            safari.implicitly_wait(2)

            try:
                safari.find_element_by_partial_link_text("百度百科")

                # 获取搜索结果页中包含"百度百科"字样的链接
                Search_Results = safari.find_elements_by_partial_link_text("百度百科")

                # 搜索结果验证
                for Search_Result in Search_Results:

                    if KeyWord in Search_Result.text:
                        continue
                    else:
                        Search_Results.remove(Search_Result)

                print(len(Search_Results))               # debug

                if len(Search_Results) > 0:
                    WS.cell(row = row_index, column = Head_Column_No + 2).value = "已收录"

                    Switch_Link = Search_Results[0].get_attribute('href')
                    print("词条链接为：%s" %(Switch_Link))                                     # Debug

                    # 打开词条页面
                    safari.get(Switch_Link)
                    safari.implicitly_wait(2)

                    WS.cell(row=row_index, column=Head_Column_No + 4).value = safari.current_url

                    # 获取词条特征数据
                    Entry_Data = Baike_Data_Scratch(safari)
                else:
                    raise NoSuchElementException
            except NoSuchElementException:
                WS.cell(row = row_index, column = Head_Column_No + 2).value = "未收录"
                WS.cell(row = row_index, column = Head_Column_No + 4).value = "None"

                Entry_Data = []
                Entry_Data.append("None")
                for index in range(1,11):
                    Entry_Data.append(-1)

        except NoSuchElementException:
            WS.cell(row = row_index, column = Head_Column_No + 2).value = "已收录"

            # 查询是否是多义词
            try:
                safari.find_element_by_class_name("lemmaWgt-subLemmaListTitle")

                Poly_Entries = safari.find_elements_by_partial_link_text(KeyWord)

                print("义项个数：%d" % (len(Poly_Entries)))

                for Entry in Poly_Entries:
                    for Key in Key_List:
                        if Key in Entry.text:
                            break
                        else:
                            continue
                    else:
                        Poly_Entries.remove(Entry)

                if len(Poly_Entries) == 1:
                    print("只有一个义项符合要求，正确")
                else:
                    print("有多个义项符合要求，错误")

                Valid_Entry_Link = Poly_Entries[0].get_attribute('href')

                print("符合要求义项链接：%s" % (Valid_Entry_Link))
                safari.get(Valid_Entry_Link)
                safari.implicitly_wait(2)

                WS.cell(row=row_index, column=Head_Column_No + 4).value = safari.current_url
            except NoSuchElementException:
                try:
                    safari.find_element_by_class_name("polysemantList-header-title")

                    Poly_Entries = safari.find_elements_by_xpath("//ul/li[@class='item']/a")
                    Poly_Entries.insert(0,safari.find_element_by_xpath("//ul/li[@class='item']/span[@class='selected']"))

                    print("义项个数：%d" % (len(Poly_Entries)))

                    for Entry in Poly_Entries:
                        for Key in Key_List:
                            if Key in Entry.text:
                                break
                            else:
                                continue
                        else:
                            Poly_Entries.remove(Entry)

                    if len(Poly_Entries) == 1:
                        print("只有一个义项符合要求，正确")
                    else:
                        print("有多个义项符合要求，错误")

                    if Poly_Entries[0].get_attribute('href'):

                        Valid_Entry_Link = Poly_Entries[0].get_attribute('href')

                        print("符合要求义项链接：%s" % (Valid_Entry_Link))
                        safari.get(Valid_Entry_Link)
                        safari.implicitly_wait(2)

                    WS.cell(row=row_index, column=Head_Column_No + 4).value = safari.current_url
                except NoSuchElementException:
                    print("本词条非多义词")

            # 获取特征点
            Entry_Data = Baike_Data_Scratch(safari)

        WS.cell(row = row_index, column = Head_Column_No + 1).value = Entry_Data[0]           # 百度百科词条名称
        WS.cell(row = row_index, column = Head_Column_No + 3).value = Entry_Data[-1]          # "科普中国百科"是否收录

        for Result_index in range(1,(len(Entry_Data) - 1)):
            WS.cell(row = row_index, column = Head_Column_No + 4 + Result_index).value = Entry_Data[Result_index]

        #time.sleep(2)

        WB.save('output_find_element.xlsx')                                   # 保存输出的Excel文件

safari.close()                                                        # 关闭safari窗口

safari.quit()                                                         # 推出safari
