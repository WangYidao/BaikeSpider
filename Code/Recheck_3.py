# 审查词条是否为多义词，如果是，提取其网址的编码

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import re
import sample

WB = load_workbook('/Users/bin/Desktop/能源利用_20161123.xlsx')

safari = webdriver.Safari()

# 主循环，遍历Worksheets

print(len(WB.sheetnames))

for ws_index in range(len(WB.sheetnames)):

    WS = WB.worksheets[ws_index]  # 打开worksheet

    Column_No = WS.max_column

    # 输入多义词数据表头
    WS.cell(row=2, column=Column_No+1).value = "是否为多义词"
    WS.cell(row=2, column=Column_No+2).value = "义项编码"

    print("表格表头共%d行，%d列" % (WS.max_row, WS.max_column))                 # debug

    # 判断词条名称所在列
    Entry_Site_Index = -1
    Baike_Included_Index = -1

    for item_index in range(1, (Column_No + 1)):

        if WS.cell(row=2, column=item_index).value == "词条网址":
            Entry_Site_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "是否被\"百度百科\"收录":
            Baike_Included_Index = item_index
        else:
            pass

    # 行循环
    for row_index in range(3, (WS.max_row + 1)):

        if WS.cell(row=row_index, column=Baike_Included_Index).value == "未收录":
            continue
        else:
            pass

            safari.get(WS.cell(row=row_index, column=Entry_Site_Index).value)

            try:
                safari.find_element_by_class_name("polysemantList-header-title")

                WS.cell(row=row_index, column=Column_No + 1).value = "是"

                History_Link = safari.find_element_by_link_text("历史版本")

                Entry_URL = History_Link.get_attribute('href')

                print(Entry_URL)

                #使用正则表达式提取最后六位数字

                MatchCode = re.search(r"[^/]\d+$",Entry_URL)

                WS.cell(row=row_index, column=Column_No + 2).value = int(MatchCode.group())

            except NoSuchElementException:
                pass

            WB.save('/Users/bin/Desktop/能源利用_20161123.xlsx')

safari.close()                                                        # 关闭safari窗口

safari.quit()                                                         # 退出safari