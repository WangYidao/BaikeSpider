# 标注"科普中国百科"已收录词条
# 审核错误词条网址是否合法年，不合法，重新抓取数据
# 审核参考文献条数是否合法，不合法，重新抓取参考文献和图片张数数据，并重新评级

from openpyxl import load_workbook                                                     # 导入Excel文件读取模块
from selenium.common.exceptions import NoSuchElementException

from selenium import webdriver

import sample

WB = load_workbook('/Users/bin/Desktop/InProgress/第一次处理/航空词表_第一次处理_20161028_2297.xlsx')                  # 打开Excel文件

safari = webdriver.Safari()

# 主循环，遍历Worksheets

for ws_index in range(len(WB.sheetnames)):

    WS = WB.worksheets[ws_index]  # 打开worksheet

    Column_No = WS.max_column

    print("表格表头共%d行，%d列" % (WS.max_row, WS.max_column))        # debug

    # 判断词条名称所在列
    Entry_Column_Index = -1
    Included_Column_Index = -1
    Entry_Site_Index = -1
    Baike_Name_Index = -1
    Baike_Included_Index = -1
    Ref_Index = -1
    Pic_Num_Index = -1

    for item_index in range(1, (Column_No + 1)):
        if WS.cell(row=2, column=item_index).value == "词条名称":
            Entry_Column_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "是否被\"科普中国百科\"收录":
            Included_Column_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "词条网址":
            Entry_Site_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "百科词条名":
            Baike_Name_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "是否被\"百度百科\"收录":
            Baike_Included_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "参考文献条数":
            Ref_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "词条图片张数":
            Pic_Num_Index = item_index
        else:
            continue

    # 标注被"科普中国百科"收录的词条
    for row_index in range(3, (WS.max_row + 1)):

        # 标注"科普中国百科"已收录词条
        if WS.cell(row=row_index, column=Included_Column_Index).value == "已收录":
            WS.cell(row=row_index, column=Included_Column_Index).font = sample.MarkFormat.Error_Font
            WS.cell(row=row_index, column=Included_Column_Index).fill = sample.MarkFormat.Error_Fill

            WS.cell(row=row_index, column=Included_Column_Index+1).value = -1
        else:
            pass

        # 审核错误词条网址
        if WS.cell(row=row_index, column=Baike_Name_Index).value == "None" and WS.cell(row=row_index, column=Entry_Site_Index).value != "None":

            # 重新抓取数据
            WS.cell(row=row_index, column=Entry_Column_Index).fill = sample.MarkFormat.Error_Fill

            KeyWord = WS.cell(row=row_index, column=Entry_Column_Index).value

            Entry_Page = sample.Get_Entry_Page(KeyWord, safari, sample.KeyList.Aero_Avia_Keys)

            if Entry_Page == -1:  # 百度百科未收录

                WS.cell(row=row_index, column=Baike_Included_Index).value = "未收录"
                WS.cell(row=row_index, column=Entry_Site_Index).value = "None"

                Entry_Data = []
                Entry_Data.append("None")
                for index in range(1, 8):
                    Entry_Data.append(-1)

                Excellent_Included = False

            else:  # 百度百科未收录
                WS.cell(row=row_index, column=Baike_Included_Index).value = "已收录"
                WS.cell(row=row_index, column=Entry_Site_Index).value = Entry_Page.current_url

                Entry_Data = sample.Data_Scratch(Entry_Page)
                Entry_Data.append(sample.Included_Label(Entry_Page))
                Excellent_Included = sample.Excellent_Label(Entry_Page)

            # 数据写入
            WS.cell(row=row_index, column=Baike_Name_Index).value = Entry_Data[0]  # 百度百科词条名称
            WS.cell(row=row_index, column=Included_Column_Index).value = Entry_Data[-1]  # "科普中国百科"是否收录

            for Result_index in range(1, (len(Entry_Data) - 1)):
                WS.cell(row=row_index, column=Entry_Site_Index+Result_index).value = Entry_Data[Result_index]

            # 词条质量评级
            # 是否为"特色词条"
            if Excellent_Included:
                print("该词条为特色词条")
                WS.cell(row=row_index, column=Entry_Column_Index).font = sample.MarkFormat.Error_Font
                WS.cell(row=row_index, column=Entry_Column_Index).fill = sample.MarkFormat.Excellent_Included_Fill

            if WS.cell(row=row_index, column=Entry_Column_Index).value != WS.cell(row=row_index,column=Baike_Name_Index).value:

                WS.cell(row=row_index, column=Baike_Name_Index).fill = sample.MarkFormat.Error_Fill
                WS.cell(row=row_index, column=Baike_Name_Index).font = sample.MarkFormat.Error_Font

            # "百度百科"未收录
            if WS.cell(row=row_index, column=Baike_Included_Index).value == "未收录":  # "百度百科"未收录的为新建词条
                WS.cell(row=row_index, column=Baike_Included_Index).font = sample.MarkFormat.Error_Font
                WS.cell(row=row_index, column=Baike_Included_Index).fill = sample.MarkFormat.Not_Included_Fill

                WS.cell(row=row_index, column=Included_Column_Index+1).value = 4  # "百度百科"未收录词条，直接置为第四等级，需新建
            elif WS.cell(row=row_index, column=Included_Column_Index).value == "已收录":  # "科普中国百科"已收录的词条不需要建设
                WS.cell(row=row_index, column=Included_Column_Index).font = sample.MarkFormat.Error_Font
                WS.cell(row=row_index, column=Included_Column_Index).fill = sample.MarkFormat.Error_Fill

                WS.cell(row=row_index, column=Included_Column_Index+1).value = -1
            else:  # 百度百科已收录而"科普中国百科"未收录
                if Entry_Data[4] == -1:  # 词条中没有正文，级别4
                    WS.cell(row=row_index, column=Included_Column_Index+1).value = 4
                elif Entry_Data[4] <= 1000:  # 正文内容过少或缺少目录，级别3
                    WS.cell(row=row_index, column=Included_Column_Index+1).value = 3
                elif Entry_Data[3] <= 0:
                    WS.cell(row=row_index, column=Included_Column_Index+1).value = 3
                elif Entry_Data[2] <= 0:  # 缺乏基本信息栏或参考文献，级别2
                    WS.cell(row=row_index, column=Included_Column_Index+1).value = 2
                elif Entry_Data[5] <= 0:
                    WS.cell(row=row_index, column=Included_Column_Index+1).value = 2
                elif Entry_Data[1] <= 0:  # 缺乏概述或图片，级别2
                    WS.cell(row=row_index, column=Included_Column_Index+1).value = 2
                elif Entry_Data[6] <= 0:
                    WS.cell(row=row_index, column=Included_Column_Index+1).value = 2
                else:  # 无结构性缺失，内容不过少，等级1
                    WS.cell(row=row_index, column=Included_Column_Index+1).value = 1

        else:
            pass

        # 重新查找123级及部分4级词条词条图片及参考文献
        if WS.cell(row=row_index, column=Baike_Included_Index).value == "已收录" and WS.cell(row=row_index, column=Included_Column_Index).value == "未收录":

            Pic_Number = WS.cell(row=row_index, column=Pic_Num_Index).value
            Ref_Number = WS.cell(row=row_index, column=Ref_Index).value

            safari.get(WS.cell(row=row_index, column=Entry_Site_Index).value)

            # 获取词条图片数量
            try:
                safari.find_element_by_link_text("更多图册")
                print("找到多个图册")
                safari.get(safari.find_element_by_link_text("更多图册").get_attribute('href'))
                safari.implicitly_wait(1)

                Pic_Number = int(safari.find_element_by_xpath("//span[@class='pic-num num']").text)

                safari.get(safari.find_element_by_class_name("return-back").get_attribute('href'))
                safari.implicitly_wait(1)
            except NoSuchElementException:
                try:
                    safari.find_element_by_class_name("summary-pic")
                    print("找到一个图册")
                    safari.get(safari.find_element_by_xpath("//div[@class='summary-pic']/a").get_attribute('href'))
                    safari.implicitly_wait(1)

                    Pic_Number = int(safari.find_element_by_xpath("//span[@style='color:#427cb8']").text)

                    safari.get(safari.find_element_by_link_text("返回词条").get_attribute('href'))
                    safari.implicitly_wait(1)
                except NoSuchElementException:
                    Pic_Number = -1

            print("词条图片数量:%d" % Pic_Number)

            if Pic_Number != WS.cell(row=row_index, column=Pic_Num_Index).value:
                WS.cell(row=row_index, column=Pic_Num_Index).value = Pic_Number
                WS.cell(row=row_index, column=Pic_Num_Index).fill = sample.MarkFormat.Error_Fill

                if WS.cell(row=row_index, column=Pic_Num_Index - 2).value == -1:  # 词条中没有正文，级别4
                    WS.cell(row=row_index, column=Included_Column_Index + 1).value = 4
                elif WS.cell(row=row_index, column=Pic_Num_Index - 2).value <= 1000:  # 正文内容过少或缺少目录，级别3
                    WS.cell(row=row_index, column=Included_Column_Index + 1).value = 3
                elif WS.cell(row=row_index, column=Pic_Num_Index - 3).value <= 0:
                    WS.cell(row=row_index, column=Included_Column_Index + 1).value = 3
                elif WS.cell(row=row_index, column=Pic_Num_Index - 4).value <= 0:  # 缺乏基本信息栏或参考文献，级别2
                    WS.cell(row=row_index, column=Included_Column_Index + 1).value = 2
                elif WS.cell(row=row_index, column=Pic_Num_Index - 1).value <= 0:
                    WS.cell(row=row_index, column=Included_Column_Index + 1).value = 2
                elif WS.cell(row=row_index, column=Pic_Num_Index - 5).value <= 0:  # 缺乏概述或图片，级别2
                    WS.cell(row=row_index, column=Included_Column_Index + 1).value = 2
                elif WS.cell(row=row_index, column=Pic_Num_Index).value <= 0:
                    WS.cell(row=row_index, column=Included_Column_Index + 1).value = 2
                else:  # 无结构性缺失，内容不过少，等级1
                    WS.cell(row=row_index, column=Included_Column_Index + 1).value = 1
            else:
                pass

            if WS.cell(row=row_index, column=Ref_Index).value <= 0:

                safari.get(WS.cell(row=row_index, column=Entry_Site_Index).value)

                # 获取参考文献条数
                try:
                    safari.find_element_by_xpath("//ul[@class='reference-list']/li[@class='reference-item ']")
                    print("找到参考文献")
                    Ref_Number = len(safari.find_elements_by_xpath("//ul[@class='reference-list']/li[@class='reference-item ']")) + len(safari.find_elements_by_xpath("//ul[@class='reference-list']/li[@class='reference-item more']"))
                except NoSuchElementException:
                    Ref_Number = -1

                print("参考文献条数:%d" % Ref_Number)

                if Ref_Number != WS.cell(row=row_index, column=Ref_Index).value:
                    WS.cell(row=row_index, column=Ref_Index).value = Ref_Number
                    WS.cell(row=row_index, column=Ref_Index).fill = sample.MarkFormat.Error_Fill

                    if WS.cell(row=row_index, column=Pic_Num_Index - 2).value == -1:  # 词条中没有正文，级别4
                        WS.cell(row=row_index, column=Included_Column_Index + 1).value = 4
                    elif WS.cell(row=row_index, column=Pic_Num_Index - 2).value <= 1000:  # 正文内容过少或缺少目录，级别3
                        WS.cell(row=row_index, column=Included_Column_Index + 1).value = 3
                    elif WS.cell(row=row_index, column=Pic_Num_Index - 3).value <= 0:
                        WS.cell(row=row_index, column=Included_Column_Index + 1).value = 3
                    elif WS.cell(row=row_index, column=Pic_Num_Index - 4).value <= 0:  # 缺乏基本信息栏或参考文献，级别2
                        WS.cell(row=row_index, column=Included_Column_Index + 1).value = 2
                    elif WS.cell(row=row_index, column=Pic_Num_Index - 1).value <= 0:
                        WS.cell(row=row_index, column=Included_Column_Index + 1).value = 2
                    elif WS.cell(row=row_index, column=Pic_Num_Index - 5).value <= 0:  # 缺乏概述或图片，级别2
                        WS.cell(row=row_index, column=Included_Column_Index + 1).value = 2
                    elif WS.cell(row=row_index, column=Pic_Num_Index).value <= 0:
                        WS.cell(row=row_index, column=Included_Column_Index + 1).value = 2
                    else:  # 无结构性缺失，内容不过少，等级1
                        WS.cell(row=row_index, column=Included_Column_Index + 1).value = 1
                else:
                    pass
            else:
                pass
        else:
            pass

        WB.save('/Users/bin/Desktop/InProgress/第一次处理/航空词表_第一次处理_20161028_2297.xlsx')

safari.close()                                                        # 关闭safari窗口

safari.quit()                                                         # 退出safari