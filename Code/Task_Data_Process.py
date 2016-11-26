# 处理百度系统后台导出的数据
# 一个垂类一个文件，一个任务一个sheet

from openpyxl import load_workbook  # 导入Excel文件读取模块
from openpyxl.styles.colors import RED, WHITE
from openpyxl.styles import PatternFill, Font
import string

WB_Before = load_workbook('/Users/bin/Desktop/任务数据/开始之前数据.xlsx')
WB_Now = load_workbook('/Users/bin/Desktop/任务数据/20161121bbck.xlsx')

WS_Before = WB_Before.active
WS_Now = WB_Now.active

print("共%d个任务，任务数据%d列" % (WS_Now.max_row - 1, WS_Now.max_column))

# 获取数据列数
Entry_Name_Now_Index = -1
Entry_Task_Now_Index = -1
Entry_Type_Now_Index = -1
Entry_State_Now_Index = -1
Editor_Now_Index = -1

Entry_Name_Before_Index = -1
Entry_Task_Before_Index = -1
Entry_Type_Before_Index = -1
Entry_State_Before_Index = -1
Editor_Before_Index = -1

for index in range(1, WS_Now.max_column + 1):

    if WS_Now.cell(row=1, column=index).value == "名称":
        Entry_Name_Now_Index = index
    elif WS_Now.cell(row=1, column=index).value == "所属任务":
        Entry_Task_Now_Index = index
    elif WS_Now.cell(row=1, column=index).value == "所属垂类":
        Entry_Type_Now_Index = index
    elif WS_Now.cell(row=1, column=index).value == "状态":
        Entry_State_Now_Index = index
    elif WS_Now.cell(row=1, column=index).value == "编辑记录":
        Editor_Now_Index = index
    else:
        pass

    if WS_Before.cell(row=1, column=index).value == "名称":
        Entry_Name_Before_Index = index
    elif WS_Before.cell(row=1, column=index).value == "所属任务":
        Entry_Task_Before_Index = index
    elif WS_Before.cell(row=1, column=index).value == "所属垂类":
        Entry_Type_Before_Index = index
    elif WS_Before.cell(row=1, column=index).value == "状态":
        Entry_State_Before_Index = index
    elif WS_Before.cell(row=1, column=index).value == "编辑记录":
        Editor_Before_Index = index
    else:
        pass

WS_Pay_Data = WB_Now.create_sheet('付款数据')

# 插入表头
WS_Pay_Data.cell(row=1, column=1).value = "词条名称"
WS_Pay_Data.cell(row=1, column=2).value = "所属任务"
WS_Pay_Data.cell(row=1, column=3).value = "所属垂类"
WS_Pay_Data.cell(row=1, column=4).value = "前状态"
WS_Pay_Data.cell(row=1, column=5).value = "后状态"
WS_Pay_Data.cell(row=1, column=6).value = "编辑者昵称"
WS_Pay_Data.cell(row=1, column=7).value = "活动词条"
WS_Pay_Data.cell(row=1, column=8).value = "付款比例"

print("Data Process Started")

for row_index in range(2, WS_Now.max_row + 1):

    # 查看是否需要处理数据
    if WS_Now.cell(row=row_index, column=Entry_State_Now_Index).value == "等待编辑" \
            or WS_Now.cell(row=row_index, column=Entry_State_Now_Index).value == "已删除":
        continue
    elif WS_Now.cell(row=row_index, column=Entry_State_Now_Index).value == "正在编辑":
        continue
    else:
        pass

    # 获取词条后数据
    Entry_Now = WS_Now.cell(row=row_index, column=Entry_Name_Now_Index).value
    Entry_Now_Task = WS_Now.cell(row=row_index, column=Entry_Task_Now_Index).value
    Entry_Now_Type = WS_Now.cell(row=row_index, column=Entry_Type_Now_Index).value
    Entry_Now_State = WS_Now.cell(row=row_index, column=Entry_State_Now_Index).value
    Entry_Now_Editor = WS_Now.cell(row=row_index, column=Editor_Now_Index).value

    # 获取当前任务词条之前的数据
    for row_before_index in range(2, WS_Before.max_row + 1):

        Entry_Before = WS_Before.cell(row=row_before_index, column=Entry_Name_Before_Index).value
        Entry_Before_Task = WS_Before.cell(row=row_before_index, column=Entry_Task_Before_Index).value

        if Entry_Now == Entry_Before and Entry_Before_Task == Entry_Now_Task:

            # 获取词条前数据
            Entry_Before_Type = WS_Before.cell(row=row_before_index, column=Entry_Type_Before_Index).value
            Entry_Before_State = WS_Before.cell(row=row_before_index, column=Entry_State_Before_Index).value
            Entry_Before_Editor = WS_Before.cell(row=row_before_index, column=Editor_Before_Index).value
            Entry_Before_Editor = Entry_Before_Editor.split(" -")[0]

            # 词条名称
            WS_Pay_Data.cell(row=(WS_Pay_Data.max_row + 1), column=1).value = Entry_Now

            # 词条所属任务
            WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=2).value = Entry_Now_Task

            # 词条所属垂类
            if Entry_Now_Type != Entry_Before_Type:
                WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=3).value = Entry_Now_Type + "/" +\
                                                                                  Entry_Before_Type
            else:
                WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=3).value = Entry_Now_Type

            # 词条前状态
            WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=4).value = Entry_Before_State

            # 词条后状态
            WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=5).value = Entry_Now_State

            # 编辑者昵称
            if str(Entry_Before_Editor) == "":
                WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=6).value = str(Entry_Now_Editor)
            elif str(Entry_Before_Editor) != str(Entry_Now_Editor):
                WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=6).value = str(Entry_Now_Editor) + "/" + str(
                    Entry_Before_Editor)
            else:
                WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=6).value = str(Entry_Now_Editor)

            # 付款比例和是否为活动词条
            In_Activity = "非"
            Pay_Percentage = 0
            if Entry_Before_State == "" or Entry_Before_State == "等待编辑":
                In_Activity = "是"

                if Entry_Now_State == "等待评审" or Entry_Now_State == "正在评审":
                    Pay_Percentage = "50%"
                elif Entry_Now_State == "已达标":
                    Pay_Percentage = "100%"
                else:
                    pass
            elif Entry_Before_State == "正在编辑":
                In_Activity = "否"

                if Entry_Now_State == "等待评审" or Entry_Now_State == "正在评审":
                    Pay_Percentage = "50%"
                elif Entry_Now_State == "已达标":
                    Pay_Percentage = "100%"
                else:
                    pass
            elif Entry_Before_State == "等待评审" or Entry_Before_State == "正在评审":
                In_Activity = "否"

                if Entry_Now_State == "已达标":
                    Pay_Percentage = "100%"
                else:
                    Pay_Percentage = "50%"
            elif Entry_Before_State == "已达标":

                In_Activity = "否"
                Pay_Percentage = "100%"
            else:
                pass

            WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=7).value = In_Activity
            WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=8).value = Pay_Percentage

            if In_Activity == "是":
                WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=7).font = Font(color=WHITE)
                WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=7).fill = PatternFill(patternType='solid',
                                                                                       start_color=RED,
                                                                                       end_color=RED)
            else:
                pass
            break
        else:
            continue
    else:
        # 表示任务词条为活动中添加的词条

        # 词条名称
        WS_Pay_Data.cell(row=(WS_Pay_Data.max_row + 1), column=1).value = Entry_Now
        # 词条所属任务
        WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=2).value = Entry_Now_Task
        # 词条所属垂类
        WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=3).value = Entry_Now_Type
        # 词条后状态
        WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=5).value = Entry_Now_State
        # 词条编辑者
        WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=6).value = Entry_Now_Editor
        # 活动词条
        WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=7).value = "是"
        WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=7).font = Font(color=WHITE)
        WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=7).fill = PatternFill(patternType='solid',
                                                                               start_color=RED,
                                                                               end_color=RED)


        if Entry_Now_State == "正在评审" or Entry_Now_State == "等待评审":
            WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=8).value = "50%"
        elif Entry_Now_State == "已达标":
            WS_Pay_Data.cell(row=WS_Pay_Data.max_row, column=8).value = "100%"
        else:
            pass

    WB_Now.save('/Users/bin/Desktop/任务数据/20161121bbck.xlsx')

    print(row_index)

print("Data Process Completed")
