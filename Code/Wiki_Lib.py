'Module to get Wikipedia entry information'

_author_ = "Yidao@Babacheku"

import time,os,string

from openpyxl.styles import PatternFill,Alignment,Font
from openpyxl.styles.colors import RED,WHITE,BLUE,DARKYELLOW
                                                                                        # 导入网页自动化测试工具selenium模块
from selenium.common.exceptions import NoSuchElementException,TimeoutException,StaleElementReferenceException         # 导入异常模块

os.environ["SELENIUM_SERVER_JAR"] = "/Users/bin/Desktop/SourceTree/BaikeEnv/selenium/selenium-server-standalone-2.48.0.jar"   # 添加selenium服务器地址环境变量

def Data_Scratch( browser ):

    # results[0]: 词条名称
    # results[1]: 概述字数
    # results[2]: 正文字数
    # results[3]: 一级目录个数
    # results[4]: 参考文献条数
    # results[5]: 图片张数

    results = []

    # 获取特征点
    # 获取百科词条名称
    try:
        print("找到词条名称")
        Original_Entry_Name = browser.find_element_by_tag_name("h1").text
        if "[" in Original_Entry_Name:
            Entry_Name = Original_Entry_Name.split("[")[0]
        else:
            Entry_Name = Original_Entry_Name
        results.append(Entry_Name)
    except NoSuchElementException:
        results.append("None")

    print("百科词条名称：%s" % results[-1])

    # 获取概述和正文内容的字数
    try:
        browser.find_element_by_css_selector("div#mw-content-text>p")
        Content_Paras = browser.find_elements_by_css_selector("div#mw-content-text>p")

        debug = Content_Paras[0].text

        results.append(len(Content_Paras[0].text))

        Number_Of_Words = 0
        for Content_Para in Content_Paras:
            Number_Of_Words += len(Content_Para.text)

        results.append(Number_Of_Words-len(Content_Paras[0].text))
    except NoSuchElementException:
        results.append(-1)
        results.append(-1)

    print("概述字数:%d" % results[-2])  # debug
    print("正文字数:%d" % results[-1])  # debug

    # 获取一级目录条数
    try:
        browser.find_element_by_tag_name("h2")
        print("找到一级目录")
        results.append(len(browser.find_elements_by_tag_name("h2")))
    except NoSuchElementException:
        results.append(-1)

    print("一级目录条数:%d" % results[-1])  # debug


    # 获取参考文献条数
    try:
        browser.find_element_by_css_selector("ol.references>li")
        print("找到参考文献")
        results.append(len(browser.find_elements_by_css_selector("ol.references>li")))
    except NoSuchElementException:
        results.append(-1)

    print("参考文献条数:%d" % results[-1])

    # 获取词条图片数量
    try:
        browser.find_element_by_xpath("//a[@class='image']")
        results.append(len(browser.find_elements_by_xpath("//a[@class='image']")))
    except NoSuchElementException:
        results.append(-1)

    print("词条图片数量:%d" % results[-1])

    return results

def Get_Entry_Page( Main_Page_URL,keyword,browser ):

    # -1: 百度百科未收录
    # 百科词条Safari句柄

    print("Python Selenium Safari Started")  # 程序开始运行提示

    while True:
        browser.get(Main_Page_URL)  # 打开网址

        try:
            browser.find_element_by_class_name("mw-wiki-logo")
            break
        except NoSuchElementException:
            continue

    browser.implicitly_wait(3)

    baike_search_key = browser.find_element_by_id("searchInput")            # 按网页元素id查找网页元素query
    baike_search_key.clear()                                                # 清除输入框里的内容
    baike_search_key.send_keys(keyword)                                     # 将获取的数据添加到输入框里

    browser.find_element_by_id("searchButton").click()                            # 单击搜索按钮

    browser.implicitly_wait(5)

    # 查询词条是否被"维基百科"收录
    while True:
        try:
            browser.find_element_by_class_name("mw-search-createlink")

            return -1
        except NoSuchElementException:

            # 查询是否是多义词
            try:
                browser.find_element_by_class_name("lmw-search-exists")

                Poly_Entries = browser.find_elements_by_partial_link_text(keyword)

                print("义项个数：%d" % (len(Poly_Entries)))

                if Poly_Entries[0].get_attribute('href'):
                    Valid_Entry_Link = Poly_Entries[0].get_attribute('href')

                    print("符合要求义项链接：%s" % (Valid_Entry_Link))
                    browser.get(Valid_Entry_Link)
                    browser.implicitly_wait(3)

                return browser
            except NoSuchElementException:
                print("本词条非多义词")
                return browser
        except TimeoutException:
            continue
        except StaleElementReferenceException:
            continue

class MarkFormat:

    # 表格格式调整
    Error_Fill = PatternFill(patternType='solid', start_color=RED, end_color=RED)
    Not_Included_Fill = PatternFill(patternType='solid', start_color=BLUE, end_color=BLUE)
    Excellent_Included_Fill = PatternFill(patternType='solid',start_color=DARKYELLOW,end_color=DARKYELLOW)
    Error_Font = Font(color=WHITE)

class KeyList:

    # 词表关键字
    Aero_Avia_Keys = ['航空', '航天', '飞行器', '飞机', '导弹', '客机', '战斗机', '轰炸机', '歼击机', '攻击机', '运输机', '直升机','无人机',
                      '火箭', '卫星', '空间站', '探测器', '飞船', '宇宙', '地球', '月球', '太阳', '深空', '火星']

    Astronomy_Keys = ['望远镜','卫星','天文学']

    Climate_Keys = ['自然','大气','气象','天气','气候','词语概念']

    Information_Keys = ['软件','程序','计算机','网络','互联网','协议','链接','代码','编程','网页','通信','通讯','传讯','磁盘']

    Energy_Keys = ['能源','材料','电池']