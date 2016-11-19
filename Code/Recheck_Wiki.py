from openpyxl import load_workbook                                                     # 导入Excel文件读取模块

from selenium import webdriver

import Wiki_Lib

WB = load_workbook('/Users/bin/Desktop/InProgress/航空航天/维基处理/维基_20161110.xlsx')                  # 打开Excel文件

Main_Page_URL_zh = "https://zh.wikipedia.org/wiki/Wikipedia:首页"
Main_Page_URL_en = "https://en.wikipedia.org/wiki/Main_Page"

safari = webdriver.Safari()

# 主循环，遍历Worksheets
for ws_index in range(len(WB.sheetnames)):

    WS = WB.worksheets[ws_index]  # 打开worksheet

    Head_Column_No = WS.max_column

    print("表格表头共%d行，%d列" % (WS.max_row, WS.max_column))  # debug

    Entry_Column_Index = -1

    # 判断词条名称所在列
    for item_index in range(1, (Head_Column_No + 1)):
        if WS.cell(row=2, column=item_index).value == "词条名称":
            Entry_Column_Index = item_index

    print("词条名称所在列为：%d" % Entry_Column_Index)

    # 表头写入
    # Head_Column_No + 1:   百科词条名
    # Head_Column_No + 2:   是否被"维基百科"收录
    # Head_Column_No + 3:   词条等级
    # Head_Column_No + 4:   词条网址
    # Head_Column_No + 5:   概述字数
    # Head_Column_No + 6:   正文字数
    # Head_Column_No + 7:   一级目录条数
    # Head_Column_No + 8:  参考文献条数
    # Head_Column_No + 9:  词条图片张数
    Sheet_Header = ["维基词条名",
                    "维基百科",
                    "词条等级",
                    "维基词条网址",
                    "概述字数",
                    "正文字数",
                    "一级目录条数",
                    "参考文献条数",
                    "词条图片张数"]

    for index in range(9):

        WS.cell(row=2, column=Head_Column_No + index + 1).value = Sheet_Header[index]

    for row_index in range(3, (WS.max_row + 1)):

        KeyWord = WS.cell(row=row_index, column=Entry_Column_Index).value

        Entry_Page = Wiki_Lib.Get_Entry_Page(Main_Page_URL_zh, KeyWord, safari)

        if Entry_Page == -1:

            WS.cell(row=row_index, column=Head_Column_No + 2).value = "未收录"
            WS.cell(row=row_index, column=Head_Column_No + 4).value = "None"

            Entry_Data = []
            for index in range(1, 6):
                Entry_Data.append(-1)

        else:

            WS.cell(row=row_index, column=Head_Column_No + 2).value = "已收录"
            WS.cell(row=row_index, column=Head_Column_No + 4).value = Entry_Page.current_url

            Entry_Data = Wiki_Lib.Data_Scratch(Entry_Page)

        # 数据写入
        WS.cell(row=row_index, column=Head_Column_No + 1).value = Entry_Data[0]  # 百度百科词条名称

        for Result_index in range(1,(len(Entry_Data))):
            WS.cell(row = row_index, column = Head_Column_No + 4 + Result_index).value = Entry_Data[Result_index]

        # 词条质量评级
        if WS.cell(row=row_index, column=Head_Column_No + 2).value == "未收录":
            WS.cell(row=row_index, column=Head_Column_No + 3).value = 4
        elif Entry_Data[2] == -1:
            WS.cell(row=row_index, column=Head_Column_No + 3).value = 4
        elif Entry_Data[2] <= 1000:
            WS.cell(row=row_index, column=Head_Column_No + 3).value = 3
        elif Entry_Data[3] <= 0:
            WS.cell(row=row_index, column=Head_Column_No + 3).value = 3
        elif Entry_Data[4] <= 0:
            WS.cell(row=row_index, column=Head_Column_No + 3).value = 2
        elif Entry_Data[5] <= 0:
            WS.cell(row=row_index, column=Head_Column_No + 3).value = 2
        elif Entry_Data[1] <= 0:
            WS.cell(row=row_index, column=Head_Column_No + 3).value = 2
        else:
            WS.cell(row=row_index, column=Head_Column_No + 3).value = 1

    WB.save('/Users/bin/Desktop/InProgress/航空航天/维基处理/维基_20161110_zh.xlsx')

safari.close()                                                        # 关闭safari窗口

safari.quit()                                                         # 退出safari