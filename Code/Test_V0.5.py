from openpyxl import load_workbook                                                     # 导入Excel文件读取模块

from selenium import webdriver

import sample

WB = load_workbook('/Users/bin/Desktop/第二次处理/核能18日夜_2.xlsx')                  # 打开Excel文件

safari = webdriver.Safari()

# 主循环，遍历Worksheets
for ws_index in range(len(WB.sheetnames)):

    WS = WB.worksheets[ws_index]  # 打开worksheet

    Head_Column_No = WS.max_column

    print("表格表头共%d行，%d列" % (WS.max_row, WS.max_column))  # debug

    # 判断词条名称所在列
    for item_index in range(1, (Head_Column_No + 1)):
        if WS.cell(row=2, column=item_index).value == "词条名称":
            Entry_Column_Index = item_index

    print("词条名称所在列为：%d" % (Entry_Column_Index))

    # 表头写入
    # Head_Column_No + 1:   百科词条名
    # Head_Column_No + 2:   是否被"百度百科"收录
    # Head_Column_No + 3:   是否被"科普中国百科"收录
    # Head_Column_No + 4:   词条等级
    # Head_Column_No + 5:   词条网址
    # Head_Column_No + 6:   概述字数
    # Head_Column_No + 7:   基本信息栏条数
    # Head_Column_No + 8:   一级目录条数
    # Head_Column_No + 9:   正文字数
    # Head_Column_No + 10:  参考文献条数
    # Head_Column_No + 11:  词条图片张数
    Sheet_Header = ["百科词条名",
                    "是否被\"百度百科\"收录",
                    "是否被\"科普中国百科\"收录",
                    "词条等级",
                    "词条网址",
                    "概述字数",
                    "基本信息栏条数",
                    "一级目录条数",
                    "正文字数",
                    "参考文献条数",
                    "词条图片张数"]

    for index in range(11):
        WS.cell(row=2, column=Head_Column_No + index + 1).value = Sheet_Header[index]

    for row_index in range(3, (WS.max_row + 1)):

        KeyWord = WS.cell(row=row_index, column=Entry_Column_Index).value                            # 获取当前sheet第i行，词条名称列数据值并打印

        Entry_Page = sample.Get_Entry_Page(KeyWord,safari,sample.KeyList.Information_Keys)                                                  # 获取词条页面

        if Entry_Page == -1:                                                                         # 百度百科未收录

            WS.cell(row=row_index, column=Head_Column_No + 2).value = "未收录"
            WS.cell(row=row_index, column=Head_Column_No + 5).value = "None"

            Entry_Data = []
            Entry_Data.append("None")
            for index in range(1, 8):
                Entry_Data.append(-1)

            Excellent_Included = False

        else:                                                                                        # 百度百科未收录
            WS.cell(row=row_index, column=Head_Column_No + 2).value = "已收录"
            WS.cell(row=row_index, column=Head_Column_No + 5).value =  Entry_Page.current_url

            Entry_Data = sample.Data_Scratch( Entry_Page )
            Entry_Data.append(sample.Included_Label(Entry_Page))
            Excellent_Included = sample.Excellent_Label(Entry_Page)

        # 数据写入
        WS.cell(row=row_index, column=Head_Column_No + 1).value = Entry_Data[0]  # 百度百科词条名称
        WS.cell(row=row_index, column=Head_Column_No + 3).value = Entry_Data[-1]  # "科普中国百科"是否收录

        for Result_index in range(1,(len(Entry_Data) - 1)):
            WS.cell(row = row_index, column = Head_Column_No + 5 + Result_index).value = Entry_Data[Result_index]

        # 词条质量评级
        # 是否为"特色词条"
        if Excellent_Included:
            print("该词条为特色词条")
            WS.cell(row=row_index, column=Entry_Column_Index).font = sample.MarkFormat.Error_Font
            WS.cell(row=row_index, column=Entry_Column_Index).fill = sample.MarkFormat.Excellent_Included_Fill

        if WS.cell(row = row_index, column = Head_Column_No + 1).value != WS.cell(row = row_index, column = Entry_Column_Index).value:
            WS.cell(row = row_index, column = Head_Column_No + 1).fill = sample.MarkFormat.Error_Fill
            WS.cell(row = row_index, column = Head_Column_No + 1).font = sample.MarkFormat.Error_Font

        # "百度百科"未收录
        if WS.cell(row = row_index, column = Head_Column_No + 2).value == "未收录":                            # "百度百科"未收录的为新建词条
            WS.cell(row = row_index, column = Head_Column_No + 2).font = sample.MarkFormat.Error_Font
            WS.cell(row = row_index, column = Head_Column_No + 2).fill = sample.MarkFormat.Not_Included_Fill

            WS.cell(row = row_index, column = Head_Column_No + 4).value = 4                                   # "百度百科"未收录词条，直接置为第四等级，需新建
        elif WS.cell(row = row_index, column = Head_Column_No + 3).value == "已收录":                          # "科普中国百科"已收录的词条不需要建设
            WS.cell(row = row_index, column = Head_Column_No + 3).font = sample.MarkFormat.Error_Font
            WS.cell(row = row_index, column = Head_Column_No + 3).fill = sample.MarkFormat.Error_Fill

            WS.cell(row = row_index, column = Head_Column_No + 4).value = -1
        else:                                                                                                 # 百度百科已收录而"科普中国百科"未收录
            if Entry_Data[4] == -1:                                                                           # 词条中没有正文，级别4
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 4
            elif Entry_Data[4] <= 1000:                                                                       # 正文内容过少或缺少目录，级别3
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 3
            elif Entry_Data[3] <= 0:
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 3
            elif Entry_Data[2] <= 0:                                                                          # 缺乏基本信息栏或参考文献，级别2
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 2
            elif Entry_Data[5] <= 0:
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 2
            elif Entry_Data[1] <= 0:                                                                          # 缺乏概述或图片，级别2
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 2
            elif Entry_Data[6] <= 0:
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 2
            else:                                                                                             # 无结构性缺失，内容不过少，等级1
                WS.cell(row=row_index, column=Head_Column_No + 4).value = 1


        WB.save('/Users/bin/Desktop/第二次处理/核能18日夜_2.xlsx')

safari.close()                                                        # 关闭safari窗口

safari.quit()                                                         # 退出safari