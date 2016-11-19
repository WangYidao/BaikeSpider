from openpyxl import Workbook

from selenium.common.exceptions import StaleElementReferenceException

from selenium import webdriver

import os,string

os.environ["SELENIUM_SERVER_JAR"] = "/Users/bin/Desktop/SourceTree/BaikeEnv/selenium/selenium-server-standalone-2.48.0.jar"   # 添加selenium服务器地址环境变量

WB = Workbook()

WS = WB.active

safari = webdriver.Safari()

safari.get("http://baike.baidu.com/usertask/view?id=202478")

Tasks_1 = safari.find_elements_by_css_selector("div.lemmaList>div")

No_Of_Task = 0

for index in range(len(Tasks_1)):

    WS.cell(row=index+1, column=1).value = Tasks_1[index].get_attribute("title")

No_Of_Task += len(Tasks_1)

for index in range(3,6):

    Page_Tags = safari.find_elements_by_css_selector('div#taskLemmaPager>a.pTag')

    Page_Tags[index].click()

    Tasks_Current = safari.find_elements_by_css_selector("div.lemmaList>div")

    for row_index in range(len(Tasks_Current)):

        WS.cell(row=No_Of_Task+1,column=index-1).value = Tasks_Current[row_index].get_attribute("title")
        No_Of_Task += 1

WB.save("Task_Retake.xlsx")

safari.close()                                                        # 关闭safari窗口

safari.quit()                                                         # 退出safari

