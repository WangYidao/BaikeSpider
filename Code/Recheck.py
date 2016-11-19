# 重新检查是否为科普中国百科收录
# 重新检查是否是特色词条

from openpyxl import load_workbook                                                     # 导入Excel文件读取模块

from selenium import webdriver

import sample

WB = load_workbook('/Users/bin/Desktop/SourceTree/Code/Test_Files/test_Poly.xlsx')                  # 打开Excel文件

safari = webdriver.Safari()

# 主循环，遍历Worksheets

for ws_index in range(len(WB.sheetnames)):

    WS = WB.worksheets[ws_index]  # 打开worksheet

    Column_No = WS.max_column

    print("表格表头共%d行，%d列" % (WS.max_row, WS.max_column))        # debug

    # 判断词条名称所在列
    Entry_Column_Index = -1
    Included_Column_Index = -1
    Entry_Site_Index = -1
    for item_index in range(1, (Column_No + 1)):
        if WS.cell(row=2, column=item_index).value == "词条名称":
            Entry_Column_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "是否被\"科普中国百科\"收录":
            Included_Column_Index = item_index
        elif WS.cell(row=2, column=item_index).value == "词条网址":
            Entry_Site_Index = item_index
        else:
            continue

    print("词条名称所在列为：%d" % Entry_Column_Index)
    print("是否被\"科普中国百科\"收录数据存储在%d列" % Included_Column_Index)
    print("词条网址所在列为%d列" % Entry_Site_Index)

    for row_index in range(3, (WS.max_row + 1)):
        if WS.cell(row=row_index,column=Included_Column_Index).value == "未收录":
            safari.get(WS.cell(row=row_index,column=Entry_Site_Index).value)

            # 重新判断是否被"科普中国百科"收录
            if sample.Included_Label(safari) == "已收录":
                WS.cell(row=row_index, column=Included_Column_Index).value = "已收录"
            else:
                pass

            # 重新判断是否是"特色词条"
            if sample.Excellent_Label(safari):
                WS.cell(row=row_index, column=Entry_Column_Index).fill = sample.MarkFormat.Excellent_Included_Fill
                WS.cell(row=row_index, column=Entry_Column_Index).font = sample.MarkFormat.Error_Font
            else:
                pass
        else:
            pass

        WB.save('/Users/bin/Desktop/SourceTree/Code/Test_Files/test_Poly.xlsx')

safari.close()                                                        # 关闭safari窗口

safari.quit()                                                         # 退出safari